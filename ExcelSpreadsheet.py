import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

products_per_supplier = {}
ttl_value_per_supplier = {}
products_under_10_inv = {}
#range by default starts with zero, and max_row isn't included, need to start from row 2 and +1 to get the last row
for product_row in range(2, product_list.max_row +1): #start from second row, first row is header
    supplier_name = product_list.cell(product_row, 4).value #give the supplier name for ea. row
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5) # creating new value into empty cell

    # calculation number of products per supplier
    if supplier_name in products_per_supplier: #check if supplier is the dictionary
        current_num_prods = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_prods + 1#if supplier already exist in dict then +1
    else:
        products_per_supplier[supplier_name] = 1

    #calculate ttl value of inventory per supplier
    if supplier_name in ttl_value_per_supplier:
        current_ttl_value = ttl_value_per_supplier.get(supplier_name)
        ttl_value_per_supplier[supplier_name] = current_ttl_value + inventory * price
    else:
        ttl_value_per_supplier[supplier_name] = inventory * price

    #logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)


    # add value for ttl inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(ttl_value_per_supplier)
print(products_under_10_inv)
inv_file.save("inventory_with_ttl_value.xlsx") #save new file with ttl value
